import pandas as pd
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os

# ─────────────────────────────────────────────
#  UTILIDADES
# ─────────────────────────────────────────────
def leer_csv(path):
    df = pd.read_csv(path, sep=';', header=0, dtype=str, encoding='latin-1')
    df.columns = df.columns.str.strip()
    df = df.dropna(how='all')
    return df

def similitud(a, b):
    return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()

def buscar_por_nombre(practica, diccionarios, umbral=0.4):
    """Busca en los 3 nomencladores por similitud de nombre. Devuelve lista de matches."""
    resultados = []
    for nombre_nom, d in diccionarios.items():
        mejor_cod = mejor_desc = ''
        mejor_score = 0.0
        for cod, desc in d.items():
            s = similitud(practica, desc)
            if s > mejor_score:
                mejor_score = s
                mejor_cod   = cod
                mejor_desc  = desc
        if mejor_score >= umbral:
            resultados.append((nombre_nom, mejor_cod, mejor_desc, mejor_score))
    return resultados

def guardar_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mapeo General'
    font_rojo = Font(color='FF0000', bold=True)
    fill_rosa = PatternFill(fill_type='solid', fgColor='FFCCCC')
    headers = df.columns.tolist()
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append(row.tolist())
        if row['Estado'] == 'No Encontrado':
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = font_rojo
                cell.fill = fill_rosa
    wb.save(path)


# ─────────────────────────────────────────────
#  VENTANA MAPEO GENERAL
# ─────────────────────────────────────────────
class VentanaMapeoGeneral(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Mapeo General de Nomencladores')
        self.resizable(False, False)

        self.f_prestador  = tk.StringVar()
        self.f_nn         = tk.StringVar()
        self.f_nn_osmiss  = tk.StringVar()
        self.f_nbu        = tk.StringVar()
        self.carpeta_salida = tk.StringVar()
        self.modo_busqueda = tk.StringVar(value='codigo')
        self.ultimo_xlsx   = None

        self._build_ui()

    def _build_ui(self):
        # ── Título ──
        tk.Label(self, text='Mapeo General de Nomencladores',
                 font=('Helvetica', 14, 'bold'), pady=12
                 ).grid(row=0, column=0, columnspan=3, sticky='ew', padx=20)

        # ── Modo de búsqueda ──
        tk.Label(self, text='Buscar por:', font=('Helvetica', 10, 'bold')
                 ).grid(row=1, column=0, sticky='w', padx=20, pady=(5, 0))

        frame_radio = tk.Frame(self)
        frame_radio.grid(row=2, column=0, columnspan=3, sticky='w', padx=30, pady=5)
        tk.Radiobutton(frame_radio, text='Código', variable=self.modo_busqueda,
                       value='codigo', font=('Helvetica', 10)
                       ).pack(side='left', padx=15)
        tk.Radiobutton(frame_radio, text='Nombre / Descripción', variable=self.modo_busqueda,
                       value='nombre', font=('Helvetica', 10)
                       ).pack(side='left', padx=15)

        # ── Archivos de entrada ──
        tk.Label(self, text='Archivos de entrada', font=('Helvetica', 10, 'bold')
                 ).grid(row=3, column=0, columnspan=3, sticky='w', padx=20, pady=(10, 0))

        archivos = [
            ('prestador.csv',  self.f_prestador),
            ('NN.csv',         self.f_nn),
            ('NN_OSMISS.csv',  self.f_nn_osmiss),
            ('NBU.csv',        self.f_nbu),
        ]
        for i, (label, var) in enumerate(archivos):
            tk.Label(self, text=label + ':'
                     ).grid(row=4+i, column=0, sticky='e', padx=(20, 5), pady=3)
            tk.Entry(self, textvariable=var, width=45, state='readonly'
                     ).grid(row=4+i, column=1, pady=3)
            tk.Button(self, text='Examinar',
                      command=lambda v=var: self._sel_archivo(v)
                      ).grid(row=4+i, column=2, padx=(5, 20), pady=3)

        fila_sal = 4 + len(archivos)

        # ── Carpeta de salida ──
        tk.Label(self, text='Carpeta de salida', font=('Helvetica', 10, 'bold')
                 ).grid(row=fila_sal, column=0, columnspan=3, sticky='w', padx=20, pady=(10, 0))
        tk.Entry(self, textvariable=self.carpeta_salida, width=45, state='readonly'
                 ).grid(row=fila_sal+1, column=1, pady=3)
        tk.Button(self, text='Examinar', command=self._sel_carpeta
                  ).grid(row=fila_sal+1, column=2, padx=(5, 20))

        # ── Progreso ──
        self.progreso = ttk.Progressbar(self, mode='indeterminate', length=400)
        self.progreso.grid(row=fila_sal+2, column=0, columnspan=3, padx=20, pady=(15, 5))

        # ── Log ──
        tk.Label(self, text='Log:').grid(row=fila_sal+3, column=0, sticky='nw', padx=20)
        self.log = tk.Text(self, height=10, width=60, state='disabled',
                           bg='#1e1e1e', fg='#d4d4d4', font=('Courier', 9))
        self.log.grid(row=fila_sal+4, column=0, columnspan=3, padx=20, pady=5)
        scroll = ttk.Scrollbar(self, command=self.log.yview)
        scroll.grid(row=fila_sal+4, column=3, sticky='ns')
        self.log['yscrollcommand'] = scroll.set

        # ── Botones ──
        frame_btn = tk.Frame(self)
        frame_btn.grid(row=fila_sal+5, column=0, columnspan=3, pady=15)

        self.btn_ejecutar = tk.Button(
            frame_btn, text='▶  Ejecutar', width=15,
            bg='#0078d4', fg='white', font=('Helvetica', 10, 'bold'),
            command=self._ejecutar)
        self.btn_ejecutar.pack(side='left', padx=10)

        self.btn_abrir = tk.Button(
            frame_btn, text='📂  Abrir carpeta', width=15,
            state='disabled', command=self._abrir_carpeta)
        self.btn_abrir.pack(side='left', padx=10)

    def _sel_archivo(self, var):
        p = filedialog.askopenfilename(filetypes=[('CSV', '*.csv'), ('Todos', '*.*')])
        if p:
            var.set(p)

    def _sel_carpeta(self):
        p = filedialog.askdirectory()
        if p:
            self.carpeta_salida.set(p)

    def _log(self, msg):
        self.log.configure(state='normal')
        self.log.insert('end', msg + '\n')
        self.log.see('end')
        self.log.configure(state='disabled')

    def _validar(self):
        for label, var in [('prestador.csv', self.f_prestador),
                            ('NN.csv', self.f_nn),
                            ('NN_OSMISS.csv', self.f_nn_osmiss),
                            ('NBU.csv', self.f_nbu)]:
            if not var.get():
                messagebox.showwarning('Falta archivo', 'Selecciona: %s' % label)
                return False
        if not self.carpeta_salida.get():
            messagebox.showwarning('Falta carpeta', 'Selecciona la carpeta de salida.')
            return False
        return True

    def _ejecutar(self):
        if not self._validar():
            return
        self.btn_ejecutar.config(state='disabled')
        self.btn_abrir.config(state='disabled')
        self.progreso.start(10)
        self.log.configure(state='normal')
        self.log.delete('1.0', 'end')
        self.log.configure(state='disabled')
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            modo = self.modo_busqueda.get()
            self._log('Modo de busqueda: %s' % ('Codigo' if modo == 'codigo' else 'Nombre'))

            self._log('Cargando archivos...')
            prestador = leer_csv(self.f_prestador.get())
            nn        = leer_csv(self.f_nn.get())
            nn_osmiss = leer_csv(self.f_nn_osmiss.get())
            nbu       = leer_csv(self.f_nbu.get())

            self._log('  prestador : %d filas' % len(prestador))

            # Limpiar columnas prestador
            prestador.columns = prestador.columns.str.strip()
            prestador['codigo_c']    = prestador['codigo_c'].str.strip()
            prestador['PRESTACIONES'] = prestador['PRESTACIONES'].str.strip()
            prestador['valor']       = prestador['valor'].str.strip()
            prestador['codigo']      = prestador['codigo'].str.strip()
            prestador['Nomenclador'] = prestador['Nomenclador'].str.strip()

            # Diccionarios nomencladores
            nn['C\xf3digo Nomenclador'] = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripci\xf3n']        = nn['Descripci\xf3n'].str.strip()
            nn_osmiss['C\xf3digo']      = nn_osmiss['C\xf3digo'].str.strip()
            nn_osmiss['Descripci\xf3n'] = nn_osmiss['Descripci\xf3n'].str.strip()
            nbu['CODIGO']              = nbu['CODIGO'].str.strip()
            nbu['Determinaciones']     = nbu['Determinaciones'].str.strip()

            dict_nn     = dict(zip(nn['C\xf3digo Nomenclador'], nn['Descripci\xf3n']))
            dict_osmiss = dict(zip(nn_osmiss['C\xf3digo'], nn_osmiss['Descripci\xf3n']))
            dict_nbu    = dict(zip(nbu['CODIGO'], nbu['Determinaciones']))

            diccionarios = {
                'NN':       dict_nn,
                'N OSMISS': dict_osmiss,
                'NBU':      dict_nbu,
            }

            self._log('Procesando %d filas...' % len(prestador))
            filas_resultado = []
            total = len(prestador)

            for i, (_, row) in enumerate(prestador.iterrows()):
                if (i+1) % 20 == 0 or (i+1) == total:
                    self._log('  Fila %d / %d' % (i+1, total))

                codigo_c   = str(row['codigo_c']).strip()
                prestacion = str(row['PRESTACIONES']).strip()
                valor      = str(row['valor']).strip()
                codigo_ref = str(row['codigo']).strip()
                nom_ref    = str(row['Nomenclador']).strip()

                encontrado = False

                if modo == 'codigo':
                    # Buscar codigo en los 3 nomencladores
                    for nombre_nom, d in diccionarios.items():
                        if codigo_ref in d:
                            filas_resultado.append({
                                'Codigo Prestador':       codigo_c,
                                'Practica Prestador':     prestacion,
                                'Valor':                  valor,
                                'Nomenclador Referencia': nom_ref,
                                'Codigo Nomenclador':     codigo_ref,
                                'Nomenclador':            nombre_nom,
                                'Descripcion Nomenclador': d[codigo_ref],
                                'Estado':                 'Encontrado',
                            })
                            encontrado = True
                            break

                else:
                    # Buscar por similitud de nombre en los 3 nomencladores
                    resultados = buscar_por_nombre(prestacion, diccionarios, umbral=0.4)
                    if resultados:
                        for nombre_nom, cod, desc, score in resultados:
                            filas_resultado.append({
                                'Codigo Prestador':       codigo_c,
                                'Practica Prestador':     prestacion,
                                'Valor':                  valor,
                                'Nomenclador Referencia': nom_ref,
                                'Codigo Nomenclador':     cod,
                                'Nomenclador':            nombre_nom,
                                'Descripcion Nomenclador': desc,
                                'Estado':                 'Encontrado',
                            })
                        encontrado = True

                if not encontrado:
                    filas_resultado.append({
                        'Codigo Prestador':       codigo_c,
                        'Practica Prestador':     prestacion,
                        'Valor':                  valor,
                        'Nomenclador Referencia': nom_ref,
                        'Codigo Nomenclador':     '',
                        'Nomenclador':            '',
                        'Descripcion Nomenclador': '',
                        'Estado':                 'No Encontrado',
                    })

            df = pd.DataFrame(filas_resultado)
            salida    = self.carpeta_salida.get()
            xlsx_path = os.path.join(salida, 'resultado_mapeo_general.xlsx')
            csv_path  = os.path.join(salida, 'resultado_mapeo_general.csv')

            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)

            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)

            self.ultimo_xlsx = xlsx_path

            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Total filas    : %d' % len(df))
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')

            self.progreso.stop()
            self.btn_ejecutar.config(state='normal')
            self.btn_abrir.config(state='normal')

        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self.progreso.stop()
            self.btn_ejecutar.config(state='normal')

    def _abrir_carpeta(self):
        if self.ultimo_xlsx:
            os.startfile(os.path.dirname(self.ultimo_xlsx))


if __name__ == '__main__':
    app = VentanaMapeoGeneral()
    app.mainloop()
