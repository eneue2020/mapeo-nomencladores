import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

UMBRAL = 0.5

def leer_csv(path):
    df = pd.read_csv(path, sep=';', header=0, dtype=str, encoding='latin-1')
    df.columns = df.columns.str.strip()
    return df.dropna(how='all')

lab      = leer_csv('laboratorios.csv')
nbu      = leer_csv('NBU.csv')
nn_osmiss = leer_csv('NN_OSMISS.csv')

lab['practica']            = lab['practica'].str.strip()
lab['valor']               = lab['valor'].str.strip()
nbu['CODIGO']              = nbu['CODIGO'].str.strip()
nbu['Determinaciones']     = nbu['Determinaciones'].str.strip()
nn_osmiss['Codigo']        = nn_osmiss['Codigo'].str.strip()
nn_osmiss['Descripcion']   = nn_osmiss['Descripcion'].str.strip()

REEMPLAZOS = {'en sangre': 'serica', 'orina': 'urinario'}
def normalizar(texto):
    t = str(texto).lower()
    for origen, destino in REEMPLAZOS.items():
        t = t.replace(origen, destino)
    return t

practicas_orig   = lab['practica'].fillna('').tolist()
practicas_norm   = [normalizar(p) for p in practicas_orig]

# --- TF-IDF contra NBU ---
desc_nbu    = nbu['Determinaciones'].fillna('').tolist()
desc_osmiss = nn_osmiss['Descripcion'].fillna('').tolist()

vec_nbu      = TfidfVectorizer().fit(desc_nbu + practicas_norm)
sim_nbu      = cosine_similarity(vec_nbu.transform(practicas_norm), vec_nbu.transform(desc_nbu))

# --- TF-IDF contra NN_OSMISS (para fallback) ---
vec_osmiss   = TfidfVectorizer().fit(desc_osmiss + practicas_norm)
sim_osmiss   = cosine_similarity(vec_osmiss.transform(practicas_norm), vec_osmiss.transform(desc_osmiss))

filas = []
for i, practica in enumerate(practicas_norm):
    if not practica:
        continue
    valor = lab.iloc[i]['valor']

    # Buscar en NBU
    indices_nbu = sorted([j for j, s in enumerate(sim_nbu[i]) if s >= UMBRAL],
                         key=lambda j: sim_nbu[i][j], reverse=True)
    if indices_nbu:
        for j in indices_nbu:
            filas.append({
                'Practica':        practicas_orig[i],
                'Valor':           valor,
                'Fuente':          'NBU',
                'Codigo':          nbu.iloc[j]['CODIGO'],
                'Descripcion':     nbu.iloc[j]['Determinaciones'],
                'Similitud':       round(sim_nbu[i][j], 4),
                'Estado':          'Encontrado',
            })
    else:
        # Fallback: buscar en NN_OSMISS
        indices_osm = sorted([j for j, s in enumerate(sim_osmiss[i]) if s >= UMBRAL],
                              key=lambda j: sim_osmiss[i][j], reverse=True)
        if indices_osm:
            for j in indices_osm:
                filas.append({
                    'Practica':    practicas_orig[i],
                    'Valor':       valor,
                    'Fuente':      'N OSMISS',
                    'Codigo':      nn_osmiss.iloc[j]['Codigo'],
                    'Descripcion': nn_osmiss.iloc[j]['Descripcion'],
                    'Similitud':   round(sim_osmiss[i][j], 4),
                    'Estado':      'Encontrado',
                })
        else:
            filas.append({
                'Practica':    practicas_orig[i],
                'Valor':       valor,
                'Fuente':      '',
                'Codigo':      '',
                'Descripcion': '',
                'Similitud':   '',
                'Estado':      'No Encontrado',
            })

df = pd.DataFrame(filas)
df.to_csv('resultado_laboratorios.csv', sep=';', index=False, encoding='latin-1')

# Excel con rojos
wb = Workbook()
ws = wb.active
ws.title = 'Laboratorios'
font_rojo = Font(color='FF0000', bold=True)
fill_rosa = PatternFill(fill_type='solid', fgColor='FFCCCC')
headers = df.columns.tolist()
ws.append(headers)
for _, row in df.iterrows():
    ws.append(row.tolist())
    if row['Estado'] == 'No Encontrado':
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = font_rojo
            cell.fill = fill_rosa
wb.save('resultado_laboratorios.xlsx')

enc = (df['Estado'] == 'Encontrado').sum()
no  = (df['Estado'] == 'No Encontrado').sum()
print(f'Encontrados:    {enc} coincidencias (umbral >= {UMBRAL})')
print(f'No encontrados: {no} practicas')
print('Archivos guardados: resultado_laboratorios.csv y resultado_laboratorios.xlsx')
