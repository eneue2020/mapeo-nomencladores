import streamlit as st
import pandas as pd
import io
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

st.set_page_config(page_title="Mapeo Nomencladores", layout="wide", page_icon="🏥")

# ─────────────────────────────────────────────
#  UTILIDADES COMUNES
# ─────────────────────────────────────────────
def leer_csv(file):
    import csv
    content = file.read()
    sample = content[:2048].decode('latin-1', errors='replace')
    sep = csv.Sniffer().sniff(sample, delimiters=';,|\t').delimiter
    import io as _io
    df = pd.read_csv(_io.BytesIO(content), sep=sep, header=0, dtype=str, encoding='latin-1', on_bad_lines='skip')
    df.columns = df.columns.str.strip()
    df = df.dropna(how='all')
    return df

def limpiar_numero(valor):
    if pd.isna(valor):
        return 0.0
    v = str(valor).strip().replace('.', '').replace(',', '.')
    try:
        return float(v)
    except ValueError:
        return 0.0

def guardar_excel_bytes(df):
    wb = Workbook()
    ws = wb.active
    font_rojo = Font(color='FF0000', bold=True)
    fill_rosa = PatternFill(fill_type='solid', fgColor='FFCCCC')
    headers = df.columns.tolist()
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append(row.tolist())
        if row['Estado'] == 'No Encontrado':
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = font_rojo
                cell.fill = fill_rosa
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def mostrar_descargas(df, nombre_base):
    if df.empty or 'Estado' not in df.columns:
        st.warning('No se generaron resultados.')
        return
    st.success('Proceso finalizado.')
    enc = (df['Estado'] == 'Encontrado').sum()
    no  = (df['Estado'] == 'No Encontrado').sum()
    col1, col2, col3 = st.columns(3)
    col1.metric('Total filas', len(df))
    col2.metric('Encontrados', enc)
    col3.metric('No encontrados', no)

    st.subheader('Descargar resultados')
    c1, c2 = st.columns(2)
    csv_bytes = df.to_csv(sep=';', index=False, encoding='latin-1').encode('latin-1')
    c1.download_button('⬇ Descargar CSV', data=csv_bytes,
                       file_name='%s.csv' % nombre_base, mime='text/csv')
    xlsx_bytes = guardar_excel_bytes(df)
    c2.download_button('⬇ Descargar Excel', data=xlsx_bytes,
                       file_name='%s.xlsx' % nombre_base,
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    st.subheader('Vista previa')
    st.dataframe(df, width='stretch')


# ─────────────────────────────────────────────
#  PROCESO 1: MAPEO CENTRO MÉDICO
# ─────────────────────────────────────────────
def proceso_mapeo():
    st.header('Mapeo Centro Médico')
    st.markdown('Cruza `rossi.csv` contra `NN.csv` y `NN_OSMISS.csv`')

    f_rossi    = st.file_uploader('rossi.csv',    type='csv', key='r_rossi')
    f_nn       = st.file_uploader('NN.csv',       type='csv', key='r_nn')
    f_nn_osmiss = st.file_uploader('NN_OSMISS.csv', type='csv', key='r_osmiss')

    if st.button('▶ Ejecutar', key='btn_mapeo'):
        if not all([f_rossi, f_nn, f_nn_osmiss]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Procesando...'):
            rossi     = leer_csv(f_rossi)
            nn        = leer_csv(f_nn)
            nn_osmiss = leer_csv(f_nn_osmiss)

            rossi['C\xf3digo']           = rossi['C\xf3digo'].str.strip()
            rossi['Nomenclador']         = rossi['Nomenclador'].str.strip()
            nn['C\xf3digo Nomenclador']  = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripci\xf3n']         = nn['Descripci\xf3n'].str.strip()
            nn_osmiss['C\xf3digo']       = nn_osmiss['C\xf3digo'].str.strip()
            nn_osmiss['Descripci\xf3n']  = nn_osmiss['Descripci\xf3n'].str.strip()

            dict_nn     = dict(zip(nn['C\xf3digo Nomenclador'], nn['Descripci\xf3n']))
            dict_osmiss = dict(zip(nn_osmiss['C\xf3digo'], nn_osmiss['Descripci\xf3n']))

            filas = []
            bar = st.progress(0, text='Procesando filas...')
            total = len(rossi)
            for i, (_, row) in enumerate(rossi.iterrows()):
                codigo      = str(row['C\xf3digo']).strip() if pd.notna(row['C\xf3digo']) else ''
                nomenclador = str(row['Nomenclador']).strip() if pd.notna(row['Nomenclador']) else ''
                if nomenclador == 'NN' and codigo in dict_nn:
                    cod_nom = codigo; desc_nom = dict_nn[codigo]; estado = 'Encontrado'
                elif nomenclador == 'N OSMISS' and codigo in dict_osmiss:
                    cod_nom = codigo; desc_nom = dict_osmiss[codigo]; estado = 'Encontrado'
                else:
                    cod_nom = desc_nom = ''; estado = 'No Encontrado'
                filas.append({
                    'C\xf3digo Rossi': row['C\xf3digo'],
                    'Nomenclador': row['Nomenclador'],
                    'Pr\xe1ctica Rossi': row.get('Pr\xe1ctica', ''),
                    'Valor': row.get('Valor', ''),
                    'Servicio': row.get('Servicio', ''),
                    'C\xf3digo Nomenclador': cod_nom,
                    'Descripci\xf3n Nomenclador': desc_nom,
                    'Estado': estado,
                })
                bar.progress(int((i+1)/total*100), text='Fila %d / %d' % (i+1, total))

            df = pd.DataFrame(filas)
        mostrar_descargas(df, 'resultado_mapeo')


# ─────────────────────────────────────────────
#  PROCESO 2: SANATORIO NN / N OSMISS
# ─────────────────────────────────────────────
def proceso_sanatorio_nn():
    st.header('Sanatorio NN / N OSMISS')
    st.markdown('Cruza `sanatorio_del_oeste.csv` contra `NN.csv` y `NN_OSMISS.csv`')

    f_san     = st.file_uploader('sanatorio_del_oeste.csv', type='csv', key='s_san')
    f_nn      = st.file_uploader('NN.csv',                  type='csv', key='s_nn')
    f_osmiss  = st.file_uploader('NN_OSMISS.csv',           type='csv', key='s_osmiss')

    if st.button('▶ Ejecutar', key='btn_san_nn'):
        if not all([f_san, f_nn, f_osmiss]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Procesando...'):
            sanatorio = leer_csv(f_san)
            nn        = leer_csv(f_nn)
            nn_osmiss = leer_csv(f_osmiss)

            nn['Codigo Nomenclador'] = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripcion']        = nn['Descripci\xf3n'].str.strip()
            nn_osmiss['Codigo']      = nn_osmiss['C\xf3digo'].str.strip()
            nn_osmiss['Descripcion'] = nn_osmiss['Descripci\xf3n'].str.strip()

            dict_nn     = dict(zip(nn['Codigo Nomenclador'], nn['Descripcion']))
            dict_osmiss = dict(zip(nn_osmiss['Codigo'], nn_osmiss['Descripcion']))

            # Agrupar
            grupos = []
            fila_principal = None
            suma_adicionales = 0.0
            for _, row in sanatorio.iterrows():
                cod  = str(row.get('codigo', '')).strip()
                nom  = str(row.get('nomenclador', '')).strip()
                copa = limpiar_numero(row.get('copago_especial', 0))
                if cod == '0' and nom in ['0', '4']:
                    suma_adicionales += copa
                else:
                    if fila_principal is not None:
                        if str(fila_principal.get('nomenclador', '')).strip() in ['2', '4']:
                            grupos.append((fila_principal, suma_adicionales))
                    fila_principal = row
                    suma_adicionales = 0.0
            if fila_principal is not None:
                if str(fila_principal.get('nomenclador', '')).strip() in ['2', '4']:
                    grupos.append((fila_principal, suma_adicionales))

            filas_resultado = []
            bar = st.progress(0, text='Procesando grupos...')
            total = len(grupos)

            for i, (row, suma_adic) in enumerate(grupos):
                codigo       = str(row.get('codigo', '')).strip()
                nom_col      = str(row.get('nomenclador', '')).strip()
                nom_col_ext  = str(row.get('NOMENCLADORES', '')).strip()
                copago_final = limpiar_numero(row.get('copago_especial', 0)) + suma_adic
                cuit         = str(row.get('cuit', '')).strip()
                nombre       = str(row.get('nombre', '')).strip()
                practicas    = str(row.get('Practicas', '')).strip()

                if codigo in ('', '0'):
                    continue

                # Expandir rango
                if ' al ' in codigo:
                    partes = codigo.split(' al ')
                    try:
                        ini = int(partes[0].strip())
                        fin = int(partes[1].strip())
                        largo = len(partes[0].strip())
                        codigos = [str(c).zfill(largo) for c in range(ini, fin+1)]
                    except ValueError:
                        codigos = [codigo]
                else:
                    codigos = [codigo]

                for cod_exp in codigos:
                    if nom_col == '2' and cod_exp in dict_nn:
                        cod_nom = cod_exp; desc_nom = dict_nn[cod_exp]; estado = 'Encontrado'
                    elif nom_col == '4' and cod_exp in dict_osmiss:
                        cod_nom = cod_exp; desc_nom = dict_osmiss[cod_exp]; estado = 'Encontrado'
                    elif cod_exp in dict_nn:
                        cod_nom = cod_exp; desc_nom = dict_nn[cod_exp]; estado = 'Encontrado'
                    elif cod_exp in dict_osmiss:
                        cod_nom = cod_exp; desc_nom = dict_osmiss[cod_exp]; estado = 'Encontrado'
                    else:
                        cod_nom = desc_nom = ''; estado = 'No Encontrado'

                    filas_resultado.append({
                        'CUIT': cuit, 'Nombre': nombre, 'Practica': practicas,
                        'Codigo': cod_exp, 'Nomenclador': nom_col_ext,
                        'Copago Especial': round(copago_final, 2),
                        'Codigo Nomenclador': cod_nom,
                        'Descripcion Nomenclador': desc_nom,
                        'Estado': estado,
                    })
                bar.progress(int((i+1)/total*100), text='Grupo %d / %d' % (i+1, total))

            df = pd.DataFrame(filas_resultado)
        mostrar_descargas(df, 'resultado_sanatorio_nn_osmiss')


# ─────────────────────────────────────────────
#  PROCESO 3: NBU
# ─────────────────────────────────────────────
def proceso_nbu():
    st.header('NBU')
    st.markdown('Cruza `sanatorio_del_oeste.csv` contra `NBU.csv`')

    f_san = st.file_uploader('sanatorio_del_oeste.csv', type='csv', key='n_san')
    f_nbu = st.file_uploader('NBU.csv',                 type='csv', key='n_nbu')

    if st.button('▶ Ejecutar', key='btn_nbu'):
        if not all([f_san, f_nbu]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Procesando...'):
            sanatorio = leer_csv(f_san)
            nbu       = leer_csv(f_nbu)

            nbu['CODIGO']          = nbu['CODIGO'].str.strip()
            nbu['Determinaciones'] = nbu['Determinaciones'].str.strip()
            dict_nbu = dict(zip(nbu['CODIGO'], nbu['Determinaciones']))

            def similitud(a, b):
                return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()

            def buscar_nbu(practica):
                mejor_cod = mejor_desc = ''
                mejor_score = 0.0
                for cod, desc in dict_nbu.items():
                    s = similitud(practica, desc)
                    if s > mejor_score:
                        mejor_score = s; mejor_cod = cod; mejor_desc = desc
                return mejor_cod, mejor_desc, mejor_score

            # Agrupar
            grupos = []
            fila_principal = None
            suma_adicionales = 0.0
            for _, row in sanatorio.iterrows():
                cod  = str(row.get('codigo', '')).strip()
                nom  = str(row.get('nomenclador', '')).strip()
                copa = limpiar_numero(row.get('copago_especial', 0))
                if cod == '0' and nom in ['0', '4']:
                    suma_adicionales += copa
                else:
                    if fila_principal is not None:
                        if str(fila_principal.get('nomenclador', '')).strip() == '3':
                            grupos.append((fila_principal, suma_adicionales))
                    fila_principal = row
                    suma_adicionales = 0.0
            if fila_principal is not None:
                if str(fila_principal.get('nomenclador', '')).strip() == '3':
                    grupos.append((fila_principal, suma_adicionales))

            UMBRAL = 0.4
            filas_resultado = []
            bar = st.progress(0, text='Procesando grupos NBU...')
            total = len(grupos)

            for i, (row, suma_adic) in enumerate(grupos):
                codigo       = str(row.get('codigo', '')).strip()
                nom_col_ext  = str(row.get('NOMENCLADORES', '')).strip()
                copago_final = limpiar_numero(row.get('copago_especial', 0)) + suma_adic
                cuit         = str(row.get('cuit', '')).strip()
                nombre       = str(row.get('nombre', '')).strip()
                practicas    = str(row.get('Practicas', '')).strip()

                if codigo not in ('', '0') and codigo in dict_nbu:
                    cod_nom = codigo; desc_nom = dict_nbu[codigo]; estado = 'Encontrado'
                else:
                    cod_nom, desc_nom, score = buscar_nbu(practicas)
                    estado = 'Encontrado por similitud' if score >= UMBRAL else 'No Encontrado'
                    if score < UMBRAL:
                        cod_nom = desc_nom = ''

                filas_resultado.append({
                    'CUIT': cuit, 'Nombre': nombre, 'Practica': practicas,
                    'Codigo': codigo if codigo != '0' else '',
                    'Nomenclador': nom_col_ext,
                    'Copago Especial': round(copago_final, 2),
                    'Codigo Nomenclador': cod_nom,
                    'Descripcion Nomenclador': desc_nom,
                    'Estado': estado,
                })
                bar.progress(int((i+1)/total*100), text='Grupo %d / %d' % (i+1, total))

            df = pd.DataFrame(filas_resultado)
        mostrar_descargas(df, 'resultado_sanatorio_nbu')


# ─────────────────────────────────────────────
#  PROCESO 4: LABORATORIOS NBU
# ─────────────────────────────────────────────
def proceso_laboratorios_nbu():
    st.header('Laboratorios NBU')
    st.markdown('Cruza `centro.csv` contra `NBU.csv` usando TF-IDF')

    f_centro = st.file_uploader('centro.csv', type='csv', key='l_centro')
    f_nbu    = st.file_uploader('NBU.csv',    type='csv', key='l_nbu')

    if st.button('▶ Ejecutar', key='btn_lab_nbu'):
        if not all([f_centro, f_nbu]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Procesando...'):
            centro = leer_csv(f_centro)
            nbu    = leer_csv(f_nbu)

            nbu['CODIGO']          = nbu['CODIGO'].str.strip()
            nbu['Determinaciones'] = nbu['Determinaciones'].str.strip()
            centro['practica']     = centro['practica'].str.strip()
            centro['valor']        = centro['valor'].str.strip()

            REEMPLAZOS = {'en sangre': 'serica', 'orina': 'urinario'}
            def normalizar(texto):
                t = str(texto).lower()
                for origen, destino in REEMPLAZOS.items():
                    t = t.replace(origen, destino)
                return t

            descripciones_nbu = nbu['Determinaciones'].fillna('').tolist()
            practicas_centro  = [normalizar(p) for p in centro['practica'].fillna('').tolist()]
            practicas_orig    = centro['practica'].fillna('').tolist()
            valores_orig      = centro['valor'].fillna('').tolist()

            st.info('Vectorizando con TF-IDF...')
            UMBRAL = 0.5
            vectorizer    = TfidfVectorizer().fit(descripciones_nbu + practicas_centro)
            matriz_nbu    = vectorizer.transform(descripciones_nbu)
            matriz_centro = vectorizer.transform(practicas_centro)
            similitudes   = cosine_similarity(matriz_centro, matriz_nbu)

            filas = []
            bar = st.progress(0, text='Procesando coincidencias...')
            total = len(practicas_centro)

            for i, practica in enumerate(practicas_centro):
                if not practica:
                    continue
                scores  = similitudes[i]
                indices = [j for j, s in enumerate(scores) if s >= UMBRAL]
                if not indices:
                    filas.append({
                        'Practica Centro': practicas_orig[i], 'Valor': valores_orig[i],
                        'COD NBU': '', 'Descripcion NBU': '',
                        'Similitud': '', 'Estado': 'No Encontrado'
                    })
                else:
                    indices.sort(key=lambda j: scores[j], reverse=True)
                    for j in indices:
                        filas.append({
                            'Practica Centro': practicas_orig[i], 'Valor': valores_orig[i],
                            'COD NBU':         nbu.iloc[j]['CODIGO'],
                            'Descripcion NBU': nbu.iloc[j]['Determinaciones'],
                            'Similitud':       round(scores[j], 4),
                            'Estado':          'Encontrado'
                        })
                bar.progress(int((i+1)/total*100), text='Practica %d / %d' % (i+1, total))

            df = pd.DataFrame(filas)
        mostrar_descargas(df, 'resultado_laboratorios_nbu')


# ─────────────────────────────────────────────
#  PROCESO 5: MAPEO GENERAL
# ─────────────────────────────────────────────
def proceso_mapeo_general():
    st.header('Mapeo General')
    st.markdown('Busca prácticas de `prestador.csv` en `NN.csv`, `NN_OSMISS.csv` y `NBU.csv`')

    modo = st.radio('Buscar por:', ['Código', 'Nombre / Descripción'], horizontal=True, key='mg_modo')

    f_prestador = st.file_uploader('prestador.csv',  type='csv', key='mg_prestador')
    f_nn        = st.file_uploader('NN.csv',         type='csv', key='mg_nn')
    f_osmiss    = st.file_uploader('NN_OSMISS.csv',  type='csv', key='mg_osmiss')
    f_nbu       = st.file_uploader('NBU.csv',        type='csv', key='mg_nbu')

    if st.button('▶ Ejecutar', key='btn_mapeo_general'):
        if not all([f_prestador, f_nn, f_osmiss, f_nbu]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Procesando...'):
            prestador = leer_csv(f_prestador)
            nn        = leer_csv(f_nn)
            nn_osmiss = leer_csv(f_osmiss)
            nbu       = leer_csv(f_nbu)

            prestador['codigo_c']     = prestador['codigo_c'].str.strip()
            prestador['PRESTACIONES'] = prestador['PRESTACIONES'].str.strip()
            prestador['valor']        = prestador['valor'].str.strip()
            prestador['codigo']       = prestador['codigo'].str.strip()
            prestador['Nomenclador']  = prestador['Nomenclador'].str.strip()

            nn['C\xf3digo Nomenclador'] = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripci\xf3n']        = nn['Descripci\xf3n'].str.strip()
            nn_osmiss['Codigo']      = nn_osmiss['Codigo'].str.strip()
            nn_osmiss['Descripcion'] = nn_osmiss['Descripcion'].str.strip()
            nbu['CODIGO']          = nbu['CODIGO'].str.strip()
            nbu['Determinaciones'] = nbu['Determinaciones'].str.strip()

            dict_nn     = {k: v for k, v in zip(nn['C\xf3digo Nomenclador'], nn['Descripci\xf3n']) if pd.notna(k) and pd.notna(v)}
            dict_osmiss = {k: v for k, v in zip(nn_osmiss['Codigo'], nn_osmiss['Descripcion']) if pd.notna(k) and pd.notna(v)}
            dict_nbu    = {k: v for k, v in zip(nbu['CODIGO'], nbu['Determinaciones']) if pd.notna(k) and pd.notna(v)}
            diccionarios = {'NN': dict_nn, 'N OSMISS': dict_osmiss, 'NBU': dict_nbu}

            def sim(a, b):
                return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()

            def buscar_nombre(practica, umbral=0.4):
                resultados = []
                for nom, d in diccionarios.items():
                    mejor_cod = mejor_desc = ''
                    mejor_score = 0.0
                    for cod, desc in d.items():
                        s = sim(practica, desc)
                        if s > mejor_score:
                            mejor_score = s; mejor_cod = cod; mejor_desc = desc
                    if mejor_score >= umbral:
                        resultados.append((nom, mejor_cod, mejor_desc))
                return resultados

            filas = []
            bar = st.progress(0, text='Procesando filas...')
            total = len(prestador)

            for i, (_, row) in enumerate(prestador.iterrows()):
                codigo_c   = str(row['codigo_c']).strip()
                prestacion = str(row['PRESTACIONES']).strip()
                valor      = str(row['valor']).strip()
                codigo_ref = str(row['codigo']).strip()
                nom_ref    = str(row['Nomenclador']).strip()
                encontrado = False

                if modo == 'Código':
                    for nom, d in diccionarios.items():
                        if codigo_ref in d:
                            filas.append({
                                'Codigo Prestador':        codigo_c,
                                'Practica Prestador':      prestacion,
                                'Valor':                   valor,
                                'Nomenclador Referencia':  nom_ref,
                                'Codigo Nomenclador':      codigo_ref,
                                'Nomenclador':             nom,
                                'Descripcion Nomenclador': d[codigo_ref],
                                'Estado':                  'Encontrado',
                            })
                            encontrado = True
                            break
                else:
                    resultados = buscar_nombre(prestacion)
                    if resultados:
                        for nom, cod, desc in resultados:
                            filas.append({
                                'Codigo Prestador':        codigo_c,
                                'Practica Prestador':      prestacion,
                                'Valor':                   valor,
                                'Nomenclador Referencia':  nom_ref,
                                'Codigo Nomenclador':      cod,
                                'Nomenclador':             nom,
                                'Descripcion Nomenclador': desc,
                                'Estado':                  'Encontrado',
                            })
                        encontrado = True

                if not encontrado:
                    filas.append({
                        'Codigo Prestador':        codigo_c,
                        'Practica Prestador':      prestacion,
                        'Valor':                   valor,
                        'Nomenclador Referencia':  nom_ref,
                        'Codigo Nomenclador':      '',
                        'Nomenclador':             '',
                        'Descripcion Nomenclador': '',
                        'Estado':                  'No Encontrado',
                    })

                bar.progress(int((i+1)/total*100), text='Fila %d / %d' % (i+1, total))

            df = pd.DataFrame(filas)
            if df.empty:
                st.warning('No se generaron resultados. Verificá los archivos cargados.')
                return
        mostrar_descargas(df, 'resultado_mapeo_general')


# ─────────────────────────────────────────────
#  PROCESO 6: LABORATORIOS
# ─────────────────────────────────────────────
def proceso_laboratorios():
    st.header('Laboratorios')
    st.markdown('Cruza `laboratorios.csv` contra `NBU.csv` y fallback a `NN_OSMISS.csv`')

    f_lab    = st.file_uploader('laboratorios.csv', type='csv', key='lab_lab')
    f_nbu    = st.file_uploader('NBU.csv',          type='csv', key='lab_nbu')
    f_osmiss = st.file_uploader('NN_OSMISS.csv',    type='csv', key='lab_osmiss')

    if st.button('▶ Ejecutar', key='btn_laboratorios'):
        if not all([f_lab, f_nbu, f_osmiss]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Procesando...'):
            lab       = leer_csv(f_lab)
            nbu       = leer_csv(f_nbu)
            nn_osmiss = leer_csv(f_osmiss)

            lab['practica']          = lab['practica'].str.strip()
            lab['valor']             = lab['valor'].str.strip()
            nbu['CODIGO']            = nbu['CODIGO'].str.strip()
            nbu['Determinaciones']   = nbu['Determinaciones'].str.strip()
            nn_osmiss['Codigo']      = nn_osmiss['Codigo'].str.strip()
            nn_osmiss['Descripcion'] = nn_osmiss['Descripcion'].str.strip()

            REEMPLAZOS = {'en sangre': 'serica', 'orina': 'urinario'}
            def normalizar(texto):
                t = str(texto).lower()
                for origen, destino in REEMPLAZOS.items():
                    t = t.replace(origen, destino)
                return t

            practicas_orig = lab['practica'].fillna('').tolist()
            practicas_norm = [normalizar(p) for p in practicas_orig]
            desc_nbu       = nbu['Determinaciones'].fillna('').tolist()
            desc_osmiss    = nn_osmiss['Descripcion'].fillna('').tolist()

            UMBRAL = 0.5
            st.info('Vectorizando con TF-IDF...')
            vec_nbu    = TfidfVectorizer().fit(desc_nbu + practicas_norm)
            sim_nbu    = cosine_similarity(vec_nbu.transform(practicas_norm), vec_nbu.transform(desc_nbu))
            vec_osmiss = TfidfVectorizer().fit(desc_osmiss + practicas_norm)
            sim_osmiss = cosine_similarity(vec_osmiss.transform(practicas_norm), vec_osmiss.transform(desc_osmiss))

            filas = []
            bar   = st.progress(0, text='Procesando coincidencias...')
            total = len(practicas_norm)

            for i, practica in enumerate(practicas_norm):
                if not practica:
                    continue
                valor = lab.iloc[i]['valor']
                indices_nbu = sorted([j for j, s in enumerate(sim_nbu[i]) if s >= UMBRAL],
                                     key=lambda j: sim_nbu[i][j], reverse=True)
                if indices_nbu:
                    for j in indices_nbu:
                        filas.append({
                            'Practica':    practicas_orig[i], 'Valor': valor,
                            'Fuente':      'NBU',
                            'Codigo':      nbu.iloc[j]['CODIGO'],
                            'Descripcion': nbu.iloc[j]['Determinaciones'],
                            'Similitud':   round(sim_nbu[i][j], 4),
                            'Estado':      'Encontrado',
                        })
                else:
                    indices_osm = sorted([j for j, s in enumerate(sim_osmiss[i]) if s >= UMBRAL],
                                         key=lambda j: sim_osmiss[i][j], reverse=True)
                    if indices_osm:
                        for j in indices_osm:
                            filas.append({
                                'Practica':    practicas_orig[i], 'Valor': valor,
                                'Fuente':      'N OSMISS',
                                'Codigo':      nn_osmiss.iloc[j]['Codigo'],
                                'Descripcion': nn_osmiss.iloc[j]['Descripcion'],
                                'Similitud':   round(sim_osmiss[i][j], 4),
                                'Estado':      'Encontrado',
                            })
                    else:
                        filas.append({
                            'Practica':    practicas_orig[i], 'Valor': valor,
                            'Fuente':      '', 'Codigo': '', 'Descripcion': '',
                            'Similitud':   '', 'Estado': 'No Encontrado',
                        })
                bar.progress(int((i+1)/total*100), text='Practica %d / %d' % (i+1, total))

            df = pd.DataFrame(filas)
        mostrar_descargas(df, 'resultado_laboratorios')


# ─────────────────────────────────────────────
#  PROCESO 7: MAPEO SEMÁNTICO (SBERT)
# ─────────────────────────────────────────────
@st.cache_resource
def cargar_modelo():
    from sentence_transformers import SentenceTransformer
    return SentenceTransformer('sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2')

def proceso_mapeo_semantico():
    st.header('Mapeo Semántico (SBERT)')
    st.markdown('Busca prácticas de `prestador.csv` en `NN.csv`, `NN_OSMISS.csv` y `NBU.csv` usando embeddings multilingües.')

    f_prestador = st.file_uploader('prestador.csv',  type='csv', key='sb_prestador')
    f_nn        = st.file_uploader('NN.csv',         type='csv', key='sb_nn')
    f_osmiss    = st.file_uploader('NN_OSMISS.csv',  type='csv', key='sb_osmiss')
    f_nbu       = st.file_uploader('NBU.csv',        type='csv', key='sb_nbu')
    top_n       = st.slider('Resultados por práctica', 1, 10, 3, key='sb_topn')

    if st.button('▶ Ejecutar', key='btn_semantico'):
        if not all([f_prestador, f_nn, f_osmiss, f_nbu]):
            st.warning('Cargá todos los archivos antes de ejecutar.')
            return

        with st.spinner('Cargando modelo SBERT (puede tardar la primera vez)...'):
            modelo = cargar_modelo()

        with st.spinner('Procesando...'):
            prestador = leer_csv(f_prestador)
            nn        = leer_csv(f_nn)
            nn_osmiss = leer_csv(f_osmiss)
            nbu       = leer_csv(f_nbu)

            prestador['PRESTACIONES'] = prestador['PRESTACIONES'].str.strip()
            nn['C\xf3digo Nomenclador'] = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripci\xf3n']        = nn['Descripci\xf3n'].str.strip()
            nn_osmiss['Codigo']        = nn_osmiss['Codigo'].str.strip()
            nn_osmiss['Descripcion']   = nn_osmiss['Descripcion'].str.strip()
            nbu['CODIGO']              = nbu['CODIGO'].str.strip()
            nbu['Determinaciones']     = nbu['Determinaciones'].str.strip()

            # Construir corpus unificado con etiqueta de fuente
            corpus = []
            for _, r in nn.iterrows():
                if pd.notna(r['Descripci\xf3n']) and pd.notna(r['C\xf3digo Nomenclador']):
                    corpus.append({'fuente': 'NN', 'codigo': r['C\xf3digo Nomenclador'], 'descripcion': r['Descripci\xf3n']})
            for _, r in nn_osmiss.iterrows():
                if pd.notna(r['Descripcion']) and pd.notna(r['Codigo']):
                    corpus.append({'fuente': 'N OSMISS', 'codigo': r['Codigo'], 'descripcion': r['Descripcion']})
            for _, r in nbu.iterrows():
                if pd.notna(r['Determinaciones']) and pd.notna(r['CODIGO']):
                    corpus.append({'fuente': 'NBU', 'codigo': r['CODIGO'], 'descripcion': r['Determinaciones']})

            textos_corpus = [c['descripcion'] for c in corpus]
            practicas     = prestador['PRESTACIONES'].fillna('').tolist()

            st.info('Generando embeddings del corpus (%d entradas)...' % len(textos_corpus))
            emb_corpus = modelo.encode(textos_corpus, batch_size=64, show_progress_bar=False, convert_to_tensor=True)

            st.info('Generando embeddings de prácticas (%d)...' % len(practicas))
            emb_practicas = modelo.encode(practicas, batch_size=64, show_progress_bar=False, convert_to_tensor=True)

            from sklearn.metrics.pairwise import cosine_similarity as cos_sim
            import numpy as np
            sims = cos_sim(emb_practicas.cpu().numpy(), emb_corpus.cpu().numpy())

            filas = []
            bar   = st.progress(0, text='Procesando resultados...')
            total = len(practicas)

            for i, practica in enumerate(practicas):
                if not practica:
                    continue
                top_idx = np.argsort(sims[i])[::-1][:top_n]
                for j in top_idx:
                    filas.append({
                        'Practica Prestador': practica,
                        'Valor':             str(prestador.iloc[i].get('valor', '')),
                        'Nomenclador':       corpus[j]['fuente'],
                        'Codigo':            corpus[j]['codigo'],
                        'Descripcion':       corpus[j]['descripcion'],
                        'Similitud':         round(float(sims[i][j]), 4),
                        'Estado':            'Encontrado' if sims[i][j] >= 0.65 else 'Baja similitud',
                    })
                bar.progress(int((i+1)/total*100), text='Práctica %d / %d' % (i+1, total))

            df = pd.DataFrame(filas)
            if df.empty:
                st.warning('No se generaron resultados.')
                return
        mostrar_descargas(df, 'resultado_mapeo_semantico')


# ─────────────────────────────────────────────
#  NAVEGACIÓN SIDEBAR
# ─────────────────────────────────────────────
st.sidebar.image('https://img.icons8.com/color/96/hospital.png', width=80)
st.sidebar.title('Mapeo de\nNomenciadores Médicos')
st.sidebar.markdown('---')

procesos = {
    '🏥 Mapeo Centro Médico':      proceso_mapeo,
    '🏨 Sanatorio NN / N OSMISS':   proceso_sanatorio_nn,
    '🧪 NBU':                      proceso_nbu,
    '🔬 Laboratorios NBU':         proceso_laboratorios_nbu,
    '🧫 Laboratorios':             proceso_laboratorios,
    '🔍 Mapeo General':            proceso_mapeo_general,
    '🧠 Mapeo Semántico (SBERT)':  proceso_mapeo_semantico,
}

seleccion = st.sidebar.radio('Seleccionar proceso:', list(procesos.keys()))
st.sidebar.markdown('---')
st.sidebar.markdown('v1.0 · [GitHub](https://github.com/eneue2020/mapeo-nomencladores)')

procesos[seleccion]()
