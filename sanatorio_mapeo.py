import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from difflib import SequenceMatcher

def leer_csv(path):
    df = pd.read_csv(path, sep=";", header=0, dtype=str, encoding="latin-1")
    df.columns = df.columns.str.strip()
    df = df.dropna(how="all")
    return df

def limpiar_numero(valor):
    if pd.isna(valor):
        return 0.0
    v = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(v)
    except ValueError:
        return 0.0

def similitud(a, b):
    return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()

def buscar_nbu(practica, dict_nbu):
    """Devuelve (codigo, descripcion, score) del mejor match en NBU."""
    mejor_cod  = ""
    mejor_desc = ""
    mejor_score = 0.0
    for cod, desc in dict_nbu.items():
        score = similitud(practica, desc)
        if score > mejor_score:
            mejor_score = score
            mejor_cod   = cod
            mejor_desc  = desc
    return mejor_cod, mejor_desc, mejor_score

def expandir_codigos(codigo, copago, cuit, nombre, nomenclador_col, nomencladores_col, practicas):
    filas = []
    if " al " in str(codigo):
        partes = str(codigo).split(" al ")
        cod_ini = partes[0].strip()
        cod_fin = partes[1].strip()
        try:
            ini   = int(cod_ini)
            fin   = int(cod_fin)
            largo = len(cod_ini)
            for cod in range(ini, fin + 1):
                filas.append({
                    "codigo": str(cod).zfill(largo), "copago_especial": copago,
                    "cuit": cuit, "nombre": nombre,
                    "nomenclador": nomenclador_col,
                    "NOMENCLADORES": nomencladores_col,
                    "Practicas": practicas,
                })
        except ValueError:
            filas.append({
                "codigo": codigo, "copago_especial": copago,
                "cuit": cuit, "nombre": nombre,
                "nomenclador": nomenclador_col,
                "NOMENCLADORES": nomencladores_col,
                "Practicas": practicas,
            })
    else:
        filas.append({
            "codigo": codigo, "copago_especial": copago,
            "cuit": cuit, "nombre": nombre,
            "nomenclador": nomenclador_col,
            "NOMENCLADORES": nomencladores_col,
            "Practicas": practicas,
        })
    return filas

# --- Cargar archivos ---
sanatorio = leer_csv("sanatorio_del_oeste.csv")
nn        = leer_csv("NN.csv")
nn_osmiss = leer_csv("NN_OSMISS.csv")
nbu       = leer_csv("NBU.csv")

# Diccionarios codigo -> descripcion
nn["C\u00f3digo Nomenclador"] = nn["C\u00f3digo Nomenclador"].str.strip()
nn["Descripci\u00f3n"]        = nn["Descripci\u00f3n"].str.strip()
nn_osmiss["C\u00f3digo"]      = nn_osmiss["C\u00f3digo"].str.strip()
nn_osmiss["Descripci\u00f3n"] = nn_osmiss["Descripci\u00f3n"].str.strip()
nbu["CODIGO"]          = nbu["CODIGO"].str.strip()
nbu["DETERMINACIONES"] = nbu["DETERMINACIONES"].str.strip()

dict_nn     = dict(zip(nn["C\u00f3digo Nomenclador"], nn["Descripci\u00f3n"]))
dict_osmiss = dict(zip(nn_osmiss["C\u00f3digo"], nn_osmiss["Descripci\u00f3n"]))
dict_nbu    = dict(zip(nbu["CODIGO"], nbu["DETERMINACIONES"]))

# --- Paso 1: agrupar filas principales con sus gastos adicionales ---
# Fila adicional: nomenclador == "0" o "4" con codigo == "0"
grupos = []
fila_principal   = None
suma_adicionales = 0.0

for _, row in sanatorio.iterrows():
    cod = str(row.get("codigo", "")).strip()
    nom = str(row.get("nomenclador", "")).strip()
    copa = limpiar_numero(row.get("copago_especial", 0))

    es_adicional = (cod == "0" and nom in ["0", "4"])

    if es_adicional:
        suma_adicionales += copa
    else:
        if fila_principal is not None:
            grupos.append((fila_principal, suma_adicionales))
        fila_principal   = row
        suma_adicionales = 0.0

if fila_principal is not None:
    grupos.append((fila_principal, suma_adicionales))

# --- Paso 2: expandir rangos y construir filas resultado ---
filas_resultado = []
UMBRAL_NBU = 0.4  # similitud minima para considerar match en NBU

for row, suma_adic in grupos:
    codigo      = str(row.get("codigo", "")).strip()
    nom_col     = str(row.get("nomenclador", "")).strip()
    nom_col_ext = str(row.get("NOMENCLADORES", "")).strip()
    copa_base   = limpiar_numero(row.get("copago_especial", 0))
    copago_final = copa_base + suma_adic
    cuit        = str(row.get("cuit", "")).strip()
    nombre      = str(row.get("nombre", "")).strip()
    practicas   = str(row.get("Practicas", "")).strip()

    if codigo == "" or codigo == "0":
        # Fila NBU sin codigo: buscar por similitud en NBU
        if nom_col == "3":
            cod_nom, desc_nom, score = buscar_nbu(practicas, dict_nbu)
            if score >= UMBRAL_NBU:
                estado = "Encontrado"
            else:
                cod_nom  = ""
                desc_nom = ""
                estado   = "No Encontrado"
            filas_resultado.append({
                "CUIT":                        cuit,
                "Nombre":                      nombre,
                "Pr\u00e1ctica":              practicas,
                "C\u00f3digo":               "",
                "Nomenclador":                nom_col_ext,
                "Copago Especial":            round(copago_final, 2),
                "C\u00f3digo Nomenclador":   cod_nom,
                "Descripci\u00f3n Nomenclador": desc_nom,
                "Estado":                     estado,
            })
        continue

    filas_exp = expandir_codigos(
        codigo, copago_final, cuit, nombre, nom_col, nom_col_ext, practicas
    )

    for f in filas_exp:
        cod_exp = str(f["codigo"]).strip()
        nom_exp = str(f["nomenclador"]).strip()

        # Buscar segun tipo de nomenclador
        if nom_exp == "2" and cod_exp in dict_nn:
            cod_nom  = cod_exp
            desc_nom = dict_nn[cod_exp]
            estado   = "Encontrado"
        elif nom_exp == "4" and cod_exp in dict_osmiss:
            cod_nom  = cod_exp
            desc_nom = dict_osmiss[cod_exp]
            estado   = "Encontrado"
        elif nom_exp == "3" and cod_exp in dict_nbu:
            cod_nom  = cod_exp
            desc_nom = dict_nbu[cod_exp]
            estado   = "Encontrado"
        elif cod_exp in dict_nn:
            cod_nom  = cod_exp
            desc_nom = dict_nn[cod_exp]
            estado   = "Encontrado"
        elif cod_exp in dict_osmiss:
            cod_nom  = cod_exp
            desc_nom = dict_osmiss[cod_exp]
            estado   = "Encontrado"
        elif cod_exp in dict_nbu:
            cod_nom  = cod_exp
            desc_nom = dict_nbu[cod_exp]
            estado   = "Encontrado"
        else:
            # Ultimo recurso: buscar por nombre en NBU si es nomenclador 3
            if nom_exp == "3":
                cod_nom, desc_nom, score = buscar_nbu(f["Practicas"], dict_nbu)
                estado = "Encontrado" if score >= UMBRAL_NBU else "No Encontrado"
                if score < UMBRAL_NBU:
                    cod_nom  = ""
                    desc_nom = ""
            else:
                cod_nom  = ""
                desc_nom = ""
                estado   = "No Encontrado"

        filas_resultado.append({
            "CUIT":                        f["cuit"],
            "Nombre":                      f["nombre"],
            "Pr\u00e1ctica":              f["Practicas"],
            "C\u00f3digo":               cod_exp,
            "Nomenclador":                f["NOMENCLADORES"],
            "Copago Especial":            round(f["copago_especial"], 2),
            "C\u00f3digo Nomenclador":   cod_nom,
            "Descripci\u00f3n Nomenclador": desc_nom,
            "Estado":                     estado,
        })

df_resultado = pd.DataFrame(filas_resultado)

# Guardar CSV
df_resultado.to_csv("resultado_sanatorio.csv", sep=";", index=False, encoding="latin-1")

# --- Generar Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Mapeo Sanatorio"

font_rojo = Font(color="FF0000", bold=True)
fill_rosa = PatternFill(fill_type="solid", fgColor="FFCCCC")

headers = df_resultado.columns.tolist()
ws.append(headers)

for _, row in df_resultado.iterrows():
    ws.append(row.tolist())
    if row["Estado"] == "No Encontrado":
        fila_excel = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=fila_excel, column=col)
            cell.font = font_rojo
            cell.fill = fill_rosa

wb.save("resultado_sanatorio.xlsx")

encontrados    = (df_resultado["Estado"] == "Encontrado").sum()
no_encontrados = (df_resultado["Estado"] == "No Encontrado").sum()
print(f"Total filas:    {len(df_resultado)}")
print(f"Encontrados:    {encontrados}")
print(f"No encontrados: {no_encontrados} (marcados en rojo/rosa)")
print("Archivos guardados: resultado_sanatorio.csv y resultado_sanatorio.xlsx")
