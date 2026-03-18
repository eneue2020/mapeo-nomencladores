import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def leer_csv(path):
    df = pd.read_csv(path, sep=";", header=0, dtype=str, encoding="latin-1")
    df.columns = df.columns.str.strip()
    df = df.dropna(how="all")
    return df

rossi     = leer_csv("rossi.csv")
nn        = leer_csv("NN.csv")
nn_osmiss = leer_csv("NN_OSMISS.csv")

# Limpiar columnas clave
rossi["C\u00f3digo"]       = rossi["C\u00f3digo"].str.strip()
rossi["Nomenclador"] = rossi["Nomenclador"].str.strip()
nn["C\u00f3digo Nomenclador"]  = nn["C\u00f3digo Nomenclador"].str.strip()
nn["Descripci\u00f3n"]         = nn["Descripci\u00f3n"].str.strip()
nn_osmiss["C\u00f3digo"]        = nn_osmiss["C\u00f3digo"].str.strip()
nn_osmiss["Descripci\u00f3n"]   = nn_osmiss["Descripci\u00f3n"].str.strip()

# Diccionarios codigo -> descripcion para lookup rapido
dict_nn     = dict(zip(nn["C\u00f3digo Nomenclador"], nn["Descripci\u00f3n"]))
dict_osmiss = dict(zip(nn_osmiss["C\u00f3digo"], nn_osmiss["Descripci\u00f3n"]))

# Construir filas del Excel
filas = []
for _, row in rossi.iterrows():
    codigo      = str(row["C\u00f3digo"]).strip() if pd.notna(row["C\u00f3digo"]) else ""
    nomenclador = str(row["Nomenclador"]).strip() if pd.notna(row["Nomenclador"]) else ""

    # Buscar descripcion segun tipo de nomenclador
    if nomenclador == "NN" and codigo in dict_nn:
        cod_nom  = codigo
        desc_nom = dict_nn[codigo]
        estado   = "Encontrado"
    elif nomenclador == "N OSMISS" and codigo in dict_osmiss:
        cod_nom  = codigo
        desc_nom = dict_osmiss[codigo]
        estado   = "Encontrado"
    else:
        cod_nom  = ""
        desc_nom = ""
        estado   = "No Encontrado"

    filas.append({
        "C\u00f3digo Rossi":        row["C\u00f3digo"],
        "Nomenclador":        row["Nomenclador"],
        "Pr\u00e1ctica Rossi":      row.get("Pr\u00e1ctica", ""),
        "Valor":              row.get("Valor", ""),
        "Servicio":           row.get("Servicio", ""),
        "C\u00f3digo Nomenclador":  cod_nom,
        "Descripci\u00f3n Nomenclador": desc_nom,
        "Estado":             estado,
    })

df_resultado = pd.DataFrame(filas)

# Guardar CSV
df_resultado.to_csv("resultado_mapeo.csv", sep=";", index=False, encoding="latin-1")

# --- Generar Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Mapeo"

font_rojo = Font(color="FF0000", bold=True)
fill_rosa = PatternFill(fill_type="solid", fgColor="FFCCCC")

headers = df_resultado.columns.tolist()
ws.append(headers)

for _, row in df_resultado.iterrows():
    ws.append(row.tolist())
    if row["Estado"] == "No Encontrado":
        fila_excel = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=fila_excel, column=col)
            cell.font = font_rojo
            cell.fill = fill_rosa

wb.save("resultado_mapeo.xlsx")

encontrados   = (df_resultado["Estado"] == "Encontrado").sum()
no_encontrados = (df_resultado["Estado"] == "No Encontrado").sum()
print(f"Encontrados:    {encontrados} registros")
print(f"No encontrados: {no_encontrados} registros (marcados en rojo/rosa)")
print("Archivos guardados: resultado_mapeo.csv y resultado_mapeo.xlsx")
