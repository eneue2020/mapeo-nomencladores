import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from difflib import SequenceMatcher
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


# ─────────────────────────────────────────────
#  UTILIDADES COMUNES
# ─────────────────────────────────────────────
def leer_csv(path):
    import csv, io
    with open(path, 'rb') as f:
        content = f.read()
    sample = content[:2048].decode('latin-1', errors='replace')
    try:
        sep = csv.Sniffer().sniff(sample, delimiters=';,|\t').delimiter
    except csv.Error:
        sep = ';'
    df = pd.read_csv(io.BytesIO(content), sep=sep, header=0, dtype=str,
                     encoding='latin-1', on_bad_lines='skip')
    df.columns = df.columns.str.strip()
    df = df.dropna(how='all')
    return df

def limpiar_numero(valor):
    if pd.isna(valor):
        return 0.0
    v = str(valor).strip().replace('.', '').replace(',', '.')
    try:
        return float(v)
    except ValueError:
        return 0.0

def guardar_excel(df, path):
    wb = Workbook()
    ws = wb.active
    font_rojo = Font(color='FF0000', bold=True)
    fill_rosa = PatternFill(fill_type='solid', fgColor='FFCCCC')
    headers = df.columns.tolist()
    ws.append(headers)
    for _, row in df.iterrows():
        ws.append(row.tolist())
        if row['Estado'] == 'No Encontrado':
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = font_rojo
                cell.fill = fill_rosa
    wb.save(path)

def abrir_carpeta(path):
    carpeta = os.path.dirname(path)
    os.startfile(carpeta)


# ─────────────────────────────────────────────
#  CLASE BASE PARA VENTANAS DE PROCESO
# ─────────────────────────────────────────────
class VentanaProceso(tk.Toplevel):
    def __init__(self, parent, titulo, archivos_requeridos):
        super().__init__(parent)
        self.title(titulo)
        self.resizable(False, False)
        self.grab_set()
        self.archivos = {}          # nombre -> StringVar con la ruta
        self.carpeta_salida = tk.StringVar()
        self.ultimo_csv  = None
        self.ultimo_xlsx = None

        self._build_ui(titulo, archivos_requeridos)

    def _build_ui(self, titulo, archivos_requeridos):
        # ── Título ──
        tk.Label(self, text=titulo, font=('Helvetica', 13, 'bold'),
                 pady=10).grid(row=0, column=0, columnspan=3, sticky='ew', padx=20)

        # ── Selectores de archivos de entrada ──
        tk.Label(self, text='Archivos de entrada', font=('Helvetica', 10, 'bold')
                 ).grid(row=1, column=0, columnspan=3, sticky='w', padx=20)

        for i, nombre in enumerate(archivos_requeridos):
            var = tk.StringVar()
            self.archivos[nombre] = var
            tk.Label(self, text=nombre + ':').grid(row=2+i, column=0, sticky='e', padx=(20, 5), pady=3)
            tk.Entry(self, textvariable=var, width=45, state='readonly'
                     ).grid(row=2+i, column=1, pady=3)
            tk.Button(self, text='Examinar',
                      command=lambda v=var: self._seleccionar_archivo(v)
                      ).grid(row=2+i, column=2, padx=(5, 20), pady=3)

        fila_salida = 2 + len(archivos_requeridos)

        # ── Carpeta de salida ──
        tk.Label(self, text='Carpeta de salida', font=('Helvetica', 10, 'bold')
                 ).grid(row=fila_salida, column=0, columnspan=3, sticky='w', padx=20, pady=(10, 0))

        tk.Entry(self, textvariable=self.carpeta_salida, width=45, state='readonly'
                 ).grid(row=fila_salida+1, column=1, pady=3)
        tk.Button(self, text='Examinar',
                  command=self._seleccionar_carpeta
                  ).grid(row=fila_salida+1, column=2, padx=(5, 20), pady=3)

        # ── Barra de progreso ──
        self.progreso = ttk.Progressbar(self, mode='indeterminate', length=400)
        self.progreso.grid(row=fila_salida+2, column=0, columnspan=3, padx=20, pady=(15, 5))

        # ── Log ──
        tk.Label(self, text='Log:').grid(row=fila_salida+3, column=0, sticky='nw', padx=20)
        self.log = tk.Text(self, height=10, width=60, state='disabled',
                           bg='#1e1e1e', fg='#d4d4d4', font=('Courier', 9))
        self.log.grid(row=fila_salida+4, column=0, columnspan=3, padx=20, pady=5)
        scroll = ttk.Scrollbar(self, command=self.log.yview)
        scroll.grid(row=fila_salida+4, column=3, sticky='ns')
        self.log['yscrollcommand'] = scroll.set

        # ── Botones de acción ──
        frame_btn = tk.Frame(self)
        frame_btn.grid(row=fila_salida+5, column=0, columnspan=3, pady=15)

        self.btn_ejecutar = tk.Button(frame_btn, text='▶  Ejecutar', width=15,
                                      bg='#0078d4', fg='white', font=('Helvetica', 10, 'bold'),
                                      command=self._ejecutar)
        self.btn_ejecutar.pack(side='left', padx=10)

        self.btn_abrir = tk.Button(frame_btn, text='📂  Abrir carpeta', width=15,
                                   state='disabled', command=self._abrir_carpeta)
        self.btn_abrir.pack(side='left', padx=10)

    def _seleccionar_archivo(self, var):
        path = filedialog.askopenfilename(filetypes=[('CSV', '*.csv'), ('Todos', '*.*')])
        if path:
            var.set(path)

    def _seleccionar_carpeta(self):
        path = filedialog.askdirectory()
        if path:
            self.carpeta_salida.set(path)

    def _log(self, msg):
        self.log.configure(state='normal')
        self.log.insert('end', msg + '\n')
        self.log.see('end')
        self.log.configure(state='disabled')

    def _validar(self):
        for nombre, var in self.archivos.items():
            if not var.get():
                messagebox.showwarning('Faltan archivos', 'Selecciona el archivo: %s' % nombre)
                return False
        if not self.carpeta_salida.get():
            messagebox.showwarning('Falta carpeta', 'Selecciona la carpeta de salida.')
            return False
        return True

    def _ejecutar(self):
        if not self._validar():
            return
        self.btn_ejecutar.config(state='disabled')
        self.btn_abrir.config(state='disabled')
        self.progreso.start(10)
        self.log.configure(state='normal')
        self.log.delete('1.0', 'end')
        self.log.configure(state='disabled')
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        raise NotImplementedError

    def _finalizar(self, ok=True):
        self.progreso.stop()
        self.btn_ejecutar.config(state='normal')
        if ok:
            self.btn_abrir.config(state='normal')

    def _abrir_carpeta(self):
        if self.ultimo_csv:
            abrir_carpeta(self.ultimo_csv)


# ─────────────────────────────────────────────
#  VENTANA: MAPEO CENTRO MÉDICO
# ─────────────────────────────────────────────
class VentanaMapeo(VentanaProceso):
    def __init__(self, parent):
        super().__init__(parent, 'Mapeo Centro Medico',
                         ['rossi.csv', 'NN.csv', 'NN_OSMISS.csv'])

    def _run(self):
        try:
            self._log('Cargando archivos...')
            rossi     = leer_csv(self.archivos['rossi.csv'].get())
            nn        = leer_csv(self.archivos['NN.csv'].get())
            nn_osmiss = leer_csv(self.archivos['NN_OSMISS.csv'].get())
            self._log('  rossi     : %d filas' % len(rossi))
            self._log('  NN        : %d filas' % len(nn))
            self._log('  NN_OSMISS : %d filas' % len(nn_osmiss))

            rossi['C\xf3digo']              = rossi['C\xf3digo'].str.strip()
            rossi['Nomenclador']           = rossi['Nomenclador'].str.strip()
            nn['C\xf3digo Nomenclador']    = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripci\xf3n']           = nn['Descripci\xf3n'].str.strip()
            dict_nn     = dict(zip(nn['C\xf3digo Nomenclador'], nn['Descripci\xf3n']))
            dict_osmiss = dict(zip(nn_osmiss['Codigo'], nn_osmiss['Descripcion']))

            self._log('Procesando filas...')
            filas = []
            for _, row in rossi.iterrows():
                codigo      = str(row['C\xf3digo']).strip() if pd.notna(row['C\xf3digo']) else ''
                nomenclador = str(row['Nomenclador']).strip() if pd.notna(row['Nomenclador']) else ''

                if nomenclador == 'NN' and codigo in dict_nn:
                    cod_nom = codigo; desc_nom = dict_nn[codigo]; estado = 'Encontrado'
                elif nomenclador == 'N OSMISS' and codigo in dict_osmiss:
                    cod_nom = codigo; desc_nom = dict_osmiss[codigo]; estado = 'Encontrado'
                else:
                    cod_nom = desc_nom = ''; estado = 'No Encontrado'

                filas.append({
                    'C\xf3digo Rossi':           row['C\xf3digo'],
                    'Nomenclador':               row['Nomenclador'],
                    'Pr\xe1ctica Rossi':         row.get('Pr\xe1ctica', ''),
                    'Valor':                     row.get('Valor', ''),
                    'Servicio':                  row.get('Servicio', ''),
                    'C\xf3digo Nomenclador':     cod_nom,
                    'Descripci\xf3n Nomenclador': desc_nom,
                    'Estado':                    estado,
                })

            df = pd.DataFrame(filas)
            salida = self.carpeta_salida.get()

            csv_path  = os.path.join(salida, 'resultado_mapeo.csv')
            xlsx_path = os.path.join(salida, 'resultado_mapeo.xlsx')

            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)

            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)

            self.ultimo_csv  = csv_path
            self.ultimo_xlsx = xlsx_path

            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')
            self._finalizar(True)

        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self._finalizar(False)


# ─────────────────────────────────────────────
#  VENTANA: SANATORIO NN / N OSMISS
# ─────────────────────────────────────────────
class VentanaSanatorioNN(VentanaProceso):
    def __init__(self, parent):
        super().__init__(parent, 'Sanatorio - NN y N OSMISS',
                         ['sanatorio_del_oeste.csv', 'NN.csv', 'NN_OSMISS.csv'])

    def _expandir(self, codigo, copago, cuit, nombre, nom_col, nom_col_ext, practicas):
        filas = []
        if ' al ' in str(codigo):
            partes = str(codigo).split(' al ')
            try:
                ini = int(partes[0].strip())
                fin = int(partes[1].strip())
                largo = len(partes[0].strip())
                for cod in range(ini, fin + 1):
                    filas.append({'codigo': str(cod).zfill(largo), 'copago_especial': copago,
                                  'cuit': cuit, 'nombre': nombre, 'nomenclador': nom_col,
                                  'NOMENCLADORES': nom_col_ext, 'Practicas': practicas})
            except ValueError:
                filas.append({'codigo': codigo, 'copago_especial': copago,
                              'cuit': cuit, 'nombre': nombre, 'nomenclador': nom_col,
                              'NOMENCLADORES': nom_col_ext, 'Practicas': practicas})
        else:
            filas.append({'codigo': codigo, 'copago_especial': copago,
                          'cuit': cuit, 'nombre': nombre, 'nomenclador': nom_col,
                          'NOMENCLADORES': nom_col_ext, 'Practicas': practicas})
        return filas

    def _run(self):
        try:
            self._log('Cargando archivos...')
            sanatorio = leer_csv(self.archivos['sanatorio_del_oeste.csv'].get())
            nn        = leer_csv(self.archivos['NN.csv'].get())
            nn_osmiss = leer_csv(self.archivos['NN_OSMISS.csv'].get())

            nn['Codigo Nomenclador'] = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripcion']        = nn['Descripci\xf3n'].str.strip()
            nn_osmiss['Codigo']      = nn_osmiss['Codigo'].str.strip()
            nn_osmiss['Descripcion'] = nn_osmiss['Descripcion'].str.strip()

            dict_nn     = dict(zip(nn['Codigo Nomenclador'], nn['Descripcion']))
            dict_osmiss = dict(zip(nn_osmiss['Codigo'], nn_osmiss['Descripcion']))
            self._log('  NN        : %d entradas' % len(dict_nn))
            self._log('  NN_OSMISS : %d entradas' % len(dict_osmiss))

            self._log('Agrupando filas...')
            grupos = []
            fila_principal = None
            suma_adicionales = 0.0

            for _, row in sanatorio.iterrows():
                cod  = str(row.get('codigo', '')).strip()
                nom  = str(row.get('nomenclador', '')).strip()
                copa = limpiar_numero(row.get('copago_especial', 0))
                es_adicional = (cod == '0' and nom in ['0', '4'])
                if es_adicional:
                    suma_adicionales += copa
                else:
                    if fila_principal is not None:
                        if str(fila_principal.get('nomenclador', '')).strip() in ['2', '4']:
                            grupos.append((fila_principal, suma_adicionales))
                    fila_principal = row
                    suma_adicionales = 0.0

            if fila_principal is not None:
                if str(fila_principal.get('nomenclador', '')).strip() in ['2', '4']:
                    grupos.append((fila_principal, suma_adicionales))

            self._log('  Grupos: %d' % len(grupos))
            self._log('Procesando grupos...')

            filas_resultado = []
            total = len(grupos)

            for i, (row, suma_adic) in enumerate(grupos):
                if (i+1) % 25 == 0 or (i+1) == total:
                    self._log('  Grupo %d / %d' % (i+1, total))

                codigo       = str(row.get('codigo', '')).strip()
                nom_col      = str(row.get('nomenclador', '')).strip()
                nom_col_ext  = str(row.get('NOMENCLADORES', '')).strip()
                copago_final = limpiar_numero(row.get('copago_especial', 0)) + suma_adic
                cuit         = str(row.get('cuit', '')).strip()
                nombre       = str(row.get('nombre', '')).strip()
                practicas    = str(row.get('Practicas', '')).strip()

                if codigo in ('', '0'):
                    continue

                for f in self._expandir(codigo, copago_final, cuit, nombre, nom_col, nom_col_ext, practicas):
                    cod_exp = str(f['codigo']).strip()
                    nom_exp = str(f['nomenclador']).strip()

                    if nom_exp == '2' and cod_exp in dict_nn:
                        cod_nom = cod_exp; desc_nom = dict_nn[cod_exp]; estado = 'Encontrado'
                    elif nom_exp == '4' and cod_exp in dict_osmiss:
                        cod_nom = cod_exp; desc_nom = dict_osmiss[cod_exp]; estado = 'Encontrado'
                    elif cod_exp in dict_nn:
                        cod_nom = cod_exp; desc_nom = dict_nn[cod_exp]; estado = 'Encontrado'
                    elif cod_exp in dict_osmiss:
                        cod_nom = cod_exp; desc_nom = dict_osmiss[cod_exp]; estado = 'Encontrado'
                    else:
                        cod_nom = desc_nom = ''; estado = 'No Encontrado'

                    filas_resultado.append({
                        'CUIT': f['cuit'], 'Nombre': f['nombre'], 'Practica': f['Practicas'],
                        'Codigo': cod_exp, 'Nomenclador': f['NOMENCLADORES'],
                        'Copago Especial': round(f['copago_especial'], 2),
                        'Codigo Nomenclador': cod_nom,
                        'Descripcion Nomenclador': desc_nom,
                        'Estado': estado,
                    })

            df = pd.DataFrame(filas_resultado)
            salida    = self.carpeta_salida.get()
            csv_path  = os.path.join(salida, 'resultado_sanatorio_nn_osmiss.csv')
            xlsx_path = os.path.join(salida, 'resultado_sanatorio_nn_osmiss.xlsx')

            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)
            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)

            self.ultimo_csv  = csv_path
            self.ultimo_xlsx = xlsx_path

            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Total filas    : %d' % len(df))
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')
            self._finalizar(True)

        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self._finalizar(False)


# ─────────────────────────────────────────────
#  VENTANA: SANATORIO NBU
# ─────────────────────────────────────────────
class VentanaNBU(VentanaProceso):
    def __init__(self, parent):
        super().__init__(parent, 'NBU',
                         ['sanatorio_del_oeste.csv', 'NBU.csv'])

    def _similitud(self, a, b):
        return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()

    def _buscar_nbu(self, practica, dict_nbu):
        mejor_cod = mejor_desc = ''
        mejor_score = 0.0
        for cod, desc in dict_nbu.items():
            score = self._similitud(practica, desc)
            if score > mejor_score:
                mejor_score = score
                mejor_cod   = cod
                mejor_desc  = desc
        return mejor_cod, mejor_desc, mejor_score

    def _run(self):
        try:
            self._log('Cargando archivos...')
            sanatorio = leer_csv(self.archivos['sanatorio_del_oeste.csv'].get())
            nbu       = leer_csv(self.archivos['NBU.csv'].get())

            nbu['CODIGO']          = nbu['CODIGO'].str.strip()
            nbu['Determinaciones'] = nbu['Determinaciones'].str.strip()
            dict_nbu = dict(zip(nbu['CODIGO'], nbu['Determinaciones']))
            self._log('  NBU: %d entradas' % len(dict_nbu))

            self._log('Agrupando filas NBU...')
            grupos = []
            fila_principal = None
            suma_adicionales = 0.0

            for _, row in sanatorio.iterrows():
                cod  = str(row.get('codigo', '')).strip()
                nom  = str(row.get('nomenclador', '')).strip()
                copa = limpiar_numero(row.get('copago_especial', 0))
                es_adicional = (cod == '0' and nom in ['0', '4'])
                if es_adicional:
                    suma_adicionales += copa
                else:
                    if fila_principal is not None:
                        if str(fila_principal.get('nomenclador', '')).strip() == '3':
                            grupos.append((fila_principal, suma_adicionales))
                    fila_principal = row
                    suma_adicionales = 0.0

            if fila_principal is not None:
                if str(fila_principal.get('nomenclador', '')).strip() == '3':
                    grupos.append((fila_principal, suma_adicionales))

            self._log('  Grupos NBU: %d' % len(grupos))
            self._log('Procesando...')

            UMBRAL = 0.4
            filas_resultado = []
            total = len(grupos)

            for i, (row, suma_adic) in enumerate(grupos):
                if (i+1) % 10 == 0 or (i+1) == total:
                    self._log('  Grupo %d / %d' % (i+1, total))

                codigo       = str(row.get('codigo', '')).strip()
                nom_col_ext  = str(row.get('NOMENCLADORES', '')).strip()
                copago_final = limpiar_numero(row.get('copago_especial', 0)) + suma_adic
                cuit         = str(row.get('cuit', '')).strip()
                nombre       = str(row.get('nombre', '')).strip()
                practicas    = str(row.get('Practicas', '')).strip()

                if codigo not in ('', '0') and codigo in dict_nbu:
                    cod_nom = codigo; desc_nom = dict_nbu[codigo]; estado = 'Encontrado'
                else:
                    cod_nom, desc_nom, score = self._buscar_nbu(practicas, dict_nbu)
                    if score >= UMBRAL:
                        estado = 'Encontrado'
                    else:
                        cod_nom = desc_nom = ''; estado = 'No Encontrado'

                filas_resultado.append({
                    'CUIT': cuit, 'Nombre': nombre, 'Practica': practicas,
                    'Codigo': codigo if codigo != '0' else '',
                    'Nomenclador': nom_col_ext,
                    'Copago Especial': round(copago_final, 2),
                    'Codigo Nomenclador': cod_nom,
                    'Descripcion Nomenclador': desc_nom,
                    'Estado': estado,
                })

            df = pd.DataFrame(filas_resultado)
            salida    = self.carpeta_salida.get()
            csv_path  = os.path.join(salida, 'resultado_sanatorio_nbu.csv')
            xlsx_path = os.path.join(salida, 'resultado_sanatorio_nbu.xlsx')

            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)
            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)

            self.ultimo_csv  = csv_path
            self.ultimo_xlsx = xlsx_path

            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Total filas    : %d' % len(df))
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')
            self._finalizar(True)

        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self._finalizar(False)


# ─────────────────────────────────────────────
#  VENTANA: LABORATORIOS NBU
# ─────────────────────────────────────────────
class VentanaLaboratoriosNBU(VentanaProceso):
    def __init__(self, parent):
        super().__init__(parent, 'Laboratorios NBU',
                         ['centro.csv', 'NBU.csv'])

    def _run(self):
        try:
            UMBRAL = 0.5
            self._log('Cargando archivos...')
            centro = leer_csv(self.archivos['centro.csv'].get())
            nbu    = leer_csv(self.archivos['NBU.csv'].get())

            col_practica = 'practica'
            col_valor    = 'valor'
            col_det      = 'Determinaciones'
            col_cod      = 'CODIGO'

            centro[col_practica] = centro[col_practica].str.strip()
            centro[col_valor]    = centro[col_valor].str.strip()
            nbu[col_det]         = nbu[col_det].str.strip()
            nbu[col_cod]         = nbu[col_cod].str.strip()

            self._log('  centro.csv : %d filas' % len(centro))
            self._log('  NBU.csv    : %d entradas' % len(nbu))

            REEMPLAZOS = {'en sangre': 'serica', 'orina': 'urinario'}
            def normalizar(texto):
                t = str(texto).lower()
                for origen, destino in REEMPLAZOS.items():
                    t = t.replace(origen, destino)
                return t

            descripciones_nbu = nbu[col_det].fillna('').tolist()
            practicas_centro  = [normalizar(p) for p in centro[col_practica].fillna('').tolist()]
            practicas_orig    = centro[col_practica].fillna('').tolist()
            valores_orig      = centro[col_valor].fillna('').tolist()

            self._log('Vectorizando con TF-IDF...')
            vectorizer    = TfidfVectorizer().fit(descripciones_nbu + practicas_centro)
            matriz_nbu    = vectorizer.transform(descripciones_nbu)
            matriz_centro = vectorizer.transform(practicas_centro)
            similitudes   = cosine_similarity(matriz_centro, matriz_nbu)

            self._log('Procesando coincidencias...')
            filas = []
            total = len(practicas_centro)
            for i, practica in enumerate(practicas_centro):
                if (i+1) % 20 == 0 or (i+1) == total:
                    self._log('  Practica %d / %d' % (i+1, total))
                if not practica:
                    continue
                scores  = similitudes[i]
                indices = [j for j, s in enumerate(scores) if s >= UMBRAL]
                if not indices:
                    filas.append({'Practica Centro': practicas_orig[i],
                                  'Valor': valores_orig[i],
                                  'COD NBU': '', 'Descripcion NBU': '',
                                  'Similitud': '', 'Estado': 'No Encontrado'})
                else:
                    indices.sort(key=lambda j: scores[j], reverse=True)
                    for j in indices:
                        filas.append({'Practica Centro': practicas_orig[i],
                                      'Valor': valores_orig[i],
                                      'COD NBU': nbu.iloc[j][col_cod],
                                      'Descripcion NBU': nbu.iloc[j][col_det],
                                      'Similitud': round(scores[j], 4),
                                      'Estado': 'Encontrado'})

            df = pd.DataFrame(filas)
            salida    = self.carpeta_salida.get()
            csv_path  = os.path.join(salida, 'resultado_laboratorios_nbu.csv')
            xlsx_path = os.path.join(salida, 'resultado_laboratorios_nbu.xlsx')

            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)
            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)

            self.ultimo_csv  = csv_path
            self.ultimo_xlsx = xlsx_path

            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')
            self._finalizar(True)

        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self._finalizar(False)


# ─────────────────────────────────────────────
#  VENTANA: LABORATORIOS
# ─────────────────────────────────────────────
class VentanaLaboratorios(VentanaProceso):
    def __init__(self, parent):
        super().__init__(parent, 'Laboratorios',
                         ['laboratorios.csv', 'NBU.csv', 'NN_OSMISS.csv'])

    def _run(self):
        try:
            UMBRAL = 0.5
            self._log('Cargando archivos...')
            lab       = leer_csv(self.archivos['laboratorios.csv'].get())
            nbu       = leer_csv(self.archivos['NBU.csv'].get())
            nn_osmiss = leer_csv(self.archivos['NN_OSMISS.csv'].get())

            lab['practica']          = lab['practica'].str.strip()
            lab['valor']             = lab['valor'].str.strip()
            nbu['CODIGO']            = nbu['CODIGO'].str.strip()
            nbu['Determinaciones']   = nbu['Determinaciones'].str.strip()
            nn_osmiss['Codigo']      = nn_osmiss['Codigo'].str.strip()
            nn_osmiss['Descripcion'] = nn_osmiss['Descripcion'].str.strip()

            self._log('  laboratorios: %d filas' % len(lab))
            self._log('  NBU         : %d entradas' % len(nbu))
            self._log('  NN_OSMISS   : %d entradas' % len(nn_osmiss))

            REEMPLAZOS = {'en sangre': 'serica', 'orina': 'urinario'}
            def normalizar(texto):
                t = str(texto).lower()
                for origen, destino in REEMPLAZOS.items():
                    t = t.replace(origen, destino)
                return t

            practicas_orig = lab['practica'].fillna('').tolist()
            practicas_norm = [normalizar(p) for p in practicas_orig]
            desc_nbu       = nbu['Determinaciones'].fillna('').tolist()
            desc_osmiss    = nn_osmiss['Descripcion'].fillna('').tolist()

            self._log('Vectorizando con TF-IDF...')
            vec_nbu    = TfidfVectorizer().fit(desc_nbu + practicas_norm)
            sim_nbu    = cosine_similarity(vec_nbu.transform(practicas_norm), vec_nbu.transform(desc_nbu))
            vec_osmiss = TfidfVectorizer().fit(desc_osmiss + practicas_norm)
            sim_osmiss = cosine_similarity(vec_osmiss.transform(practicas_norm), vec_osmiss.transform(desc_osmiss))

            self._log('Procesando coincidencias...')
            filas = []
            total = len(practicas_norm)
            for i, practica in enumerate(practicas_norm):
                if (i+1) % 20 == 0 or (i+1) == total:
                    self._log('  Practica %d / %d' % (i+1, total))
                if not practica:
                    continue
                valor = lab.iloc[i]['valor']

                indices_nbu = sorted([j for j, s in enumerate(sim_nbu[i]) if s >= UMBRAL],
                                     key=lambda j: sim_nbu[i][j], reverse=True)
                if indices_nbu:
                    for j in indices_nbu:
                        filas.append({
                            'Practica':    practicas_orig[i], 'Valor': valor,
                            'Fuente':      'NBU',
                            'Codigo':      nbu.iloc[j]['CODIGO'],
                            'Descripcion': nbu.iloc[j]['Determinaciones'],
                            'Similitud':   round(sim_nbu[i][j], 4),
                            'Estado':      'Encontrado',
                        })
                else:
                    indices_osm = sorted([j for j, s in enumerate(sim_osmiss[i]) if s >= UMBRAL],
                                         key=lambda j: sim_osmiss[i][j], reverse=True)
                    if indices_osm:
                        for j in indices_osm:
                            filas.append({
                                'Practica':    practicas_orig[i], 'Valor': valor,
                                'Fuente':      'N OSMISS',
                                'Codigo':      nn_osmiss.iloc[j]['Codigo'],
                                'Descripcion': nn_osmiss.iloc[j]['Descripcion'],
                                'Similitud':   round(sim_osmiss[i][j], 4),
                                'Estado':      'Encontrado',
                            })
                    else:
                        filas.append({
                            'Practica':    practicas_orig[i], 'Valor': valor,
                            'Fuente':      '', 'Codigo': '', 'Descripcion': '',
                            'Similitud':   '', 'Estado': 'No Encontrado',
                        })

            df = pd.DataFrame(filas)
            salida    = self.carpeta_salida.get()
            csv_path  = os.path.join(salida, 'resultado_laboratorios.csv')
            xlsx_path = os.path.join(salida, 'resultado_laboratorios.xlsx')
            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)
            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)
            self.ultimo_csv  = csv_path
            self.ultimo_xlsx = xlsx_path
            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')
            self._finalizar(True)
        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self._finalizar(False)


# ─────────────────────────────────────────────
#  VENTANA: MAPEO GENERAL
# ─────────────────────────────────────────────
class VentanaMapeoGeneral(VentanaProceso):
    def __init__(self, parent):
        self.modo = tk.StringVar(value='Código')
        super().__init__(parent, 'Mapeo General',
                         ['prestador.csv', 'NN.csv', 'NN_OSMISS.csv', 'NBU.csv'])

    def _build_ui(self, titulo, archivos_requeridos):
        super()._build_ui(titulo, archivos_requeridos)
        # Insertar radio buttons después del título
        frame_modo = tk.Frame(self)
        frame_modo.grid(row=1, column=0, columnspan=3, sticky='w', padx=20, pady=(0, 5))
        tk.Label(frame_modo, text='Buscar por:').pack(side='left')
        tk.Radiobutton(frame_modo, text='Código',             variable=self.modo, value='Código').pack(side='left', padx=5)
        tk.Radiobutton(frame_modo, text='Nombre/Descripción', variable=self.modo, value='Nombre').pack(side='left', padx=5)

    def _run(self):
        try:
            self._log('Cargando archivos...')
            prestador = leer_csv(self.archivos['prestador.csv'].get())
            nn        = leer_csv(self.archivos['NN.csv'].get())
            nn_osmiss = leer_csv(self.archivos['NN_OSMISS.csv'].get())
            nbu       = leer_csv(self.archivos['NBU.csv'].get())

            prestador['codigo_c']     = prestador['codigo_c'].str.strip()
            prestador['PRESTACIONES'] = prestador['PRESTACIONES'].str.strip()
            prestador['valor']        = prestador['valor'].str.strip()
            prestador['codigo']       = prestador['codigo'].str.strip()
            prestador['Nomenclador']  = prestador['Nomenclador'].str.strip()

            nn['C\xf3digo Nomenclador'] = nn['C\xf3digo Nomenclador'].str.strip()
            nn['Descripci\xf3n']        = nn['Descripci\xf3n'].str.strip()
            nbu['CODIGO']          = nbu['CODIGO'].str.strip()
            nbu['Determinaciones'] = nbu['Determinaciones'].str.strip()

            dict_nn     = dict(zip(nn['C\xf3digo Nomenclador'], nn['Descripci\xf3n']))
            dict_osmiss = dict(zip(nn_osmiss['Codigo'], nn_osmiss['Descripcion']))
            dict_nbu    = dict(zip(nbu['CODIGO'], nbu['Determinaciones']))
            diccionarios = {'NN': dict_nn, 'N OSMISS': dict_osmiss, 'NBU': dict_nbu}

            def sim(a, b):
                return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()

            def buscar_nombre(practica, umbral=0.4):
                resultados = []
                for nom, d in diccionarios.items():
                    mejor_cod = mejor_desc = ''
                    mejor_score = 0.0
                    for cod, desc in d.items():
                        s = sim(practica, desc)
                        if s > mejor_score:
                            mejor_score = s; mejor_cod = cod; mejor_desc = desc
                    if mejor_score >= umbral:
                        resultados.append((nom, mejor_cod, mejor_desc))
                return resultados

            modo   = self.modo.get()
            filas  = []
            total  = len(prestador)
            self._log('Procesando %d filas (modo: %s)...' % (total, modo))

            for i, (_, row) in enumerate(prestador.iterrows()):
                if (i+1) % 20 == 0 or (i+1) == total:
                    self._log('  Fila %d / %d' % (i+1, total))

                codigo_c   = str(row['codigo_c']).strip()
                prestacion = str(row['PRESTACIONES']).strip()
                valor      = str(row['valor']).strip()
                codigo_ref = str(row['codigo']).strip()
                nom_ref    = str(row['Nomenclador']).strip()
                encontrado = False

                if modo == 'Código':
                    for nom, d in diccionarios.items():
                        if codigo_ref in d:
                            filas.append({
                                'Codigo Prestador':        codigo_c,
                                'Practica Prestador':      prestacion,
                                'Valor':                   valor,
                                'Nomenclador Referencia':  nom_ref,
                                'Codigo Nomenclador':      codigo_ref,
                                'Nomenclador':             nom,
                                'Descripcion Nomenclador': d[codigo_ref],
                                'Estado':                  'Encontrado',
                            })
                            encontrado = True
                            break
                else:
                    resultados = buscar_nombre(prestacion)
                    if resultados:
                        for nom, cod, desc in resultados:
                            filas.append({
                                'Codigo Prestador':        codigo_c,
                                'Practica Prestador':      prestacion,
                                'Valor':                   valor,
                                'Nomenclador Referencia':  nom_ref,
                                'Codigo Nomenclador':      cod,
                                'Nomenclador':             nom,
                                'Descripcion Nomenclador': desc,
                                'Estado':                  'Encontrado',
                            })
                        encontrado = True

                if not encontrado:
                    filas.append({
                        'Codigo Prestador':        codigo_c,
                        'Practica Prestador':      prestacion,
                        'Valor':                   valor,
                        'Nomenclador Referencia':  nom_ref,
                        'Codigo Nomenclador':      '',
                        'Nomenclador':             '',
                        'Descripcion Nomenclador': '',
                        'Estado':                  'No Encontrado',
                    })

            df = pd.DataFrame(filas)
            salida    = self.carpeta_salida.get()
            csv_path  = os.path.join(salida, 'resultado_mapeo_general.csv')
            xlsx_path = os.path.join(salida, 'resultado_mapeo_general.xlsx')

            df.to_csv(csv_path, sep=';', index=False, encoding='latin-1')
            self._log('CSV guardado: %s' % csv_path)
            guardar_excel(df, xlsx_path)
            self._log('Excel guardado: %s' % xlsx_path)

            self.ultimo_csv  = csv_path
            self.ultimo_xlsx = xlsx_path

            enc = (df['Estado'] == 'Encontrado').sum()
            no  = (df['Estado'] == 'No Encontrado').sum()
            self._log('\n=== RESUMEN ===')
            self._log('  Total filas    : %d' % len(df))
            self._log('  Encontrados    : %d' % enc)
            self._log('  No encontrados : %d' % no)
            self._log('Proceso finalizado.')
            self._finalizar(True)

        except Exception as e:
            self._log('ERROR: %s' % str(e))
            self._finalizar(False)


# ─────────────────────────────────────────────
#  VENTANA PRINCIPAL
# ─────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Mapeo de Nomencladores Medicos')
        self.resizable(False, False)
        self._build_ui()

    def _build_ui(self):
        tk.Label(self, text='Mapeo de Nomencladores Medicos',
                 font=('Helvetica', 15, 'bold'), pady=20
                 ).pack()

        tk.Label(self, text='Selecciona el proceso que deseas ejecutar:',
                 font=('Helvetica', 10), pady=5
                 ).pack()

        frame = tk.Frame(self, pady=20)
        frame.pack()

        botones = [
            ('Mapeo Centro Medico',     VentanaMapeo),
            ('Sanatorio NN / N OSMISS', VentanaSanatorioNN),
            ('NBU',                     VentanaNBU),
            ('Laboratorios NBU',        VentanaLaboratoriosNBU),
            ('Laboratorios',            VentanaLaboratorios),
            ('Mapeo General',           VentanaMapeoGeneral),
        ]

        for texto, clase in botones:
            tk.Button(
                frame, text=texto, width=28,
                font=('Helvetica', 11), pady=8,
                bg='#0078d4', fg='white',
                command=lambda c=clase: c(self)
            ).pack(pady=6)

        tk.Label(self, text='v1.0', font=('Helvetica', 8), fg='gray', pady=10).pack()


if __name__ == '__main__':
    app = App()
    app.mainloop()
