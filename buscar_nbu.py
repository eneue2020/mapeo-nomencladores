import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

UMBRAL = 0.5

def leer_csv(path):
    df = pd.read_csv(path, sep=";", header=0, dtype=str, encoding="latin-1")
    df.columns = df.columns.str.strip()
    return df.dropna(how="all")

centro = leer_csv("centro.csv")
nbu    = leer_csv("NBU.csv")

centro["practica"] = centro["practica"].str.strip()
centro["Valor"]    = centro["valor"].str.strip()
nbu["Determinaciones"] = nbu["Determinaciones"].str.strip()
nbu["CODIGO"] = nbu["CODIGO"].str.strip()

descripciones_nbu = nbu["Determinaciones"].fillna("").tolist()
REEMPLAZOS = {
    "en sangre": "sérica",
    "orina":     "urinario",
}

def normalizar(texto):
    t = texto.lower()
    for origen, destino in REEMPLAZOS.items():
        t = t.replace(origen, destino)
    return t

practicas_centro  = [normalizar(p) for p in centro["practica"].fillna("").tolist()]
practicas_orig    = centro["practica"].fillna("").tolist()

# Vectorizar todas las descripciones juntas para espacio TF-IDF compartido
vectorizer = TfidfVectorizer().fit(descripciones_nbu + practicas_centro)
matriz_nbu    = vectorizer.transform(descripciones_nbu)
matriz_centro = vectorizer.transform(practicas_centro)

# Calcular similitud de todas las practicas contra todas las descripciones NBU
similitudes = cosine_similarity(matriz_centro, matriz_nbu)

filas = []
for i, practica in enumerate(practicas_centro):
    practica_orig = practicas_orig[i]
    if not practica:
        continue
    scores = similitudes[i]
    indices_match = [j for j, s in enumerate(scores) if s >= UMBRAL]

    if not indices_match:
        filas.append({"Practica Centro": practica_orig, "Valor": centro.iloc[i]["valor"], "COD NBU": "", "Descripcion NBU": "", "Similitud": "", "Estado": "No Encontrado"})
    else:
        # Ordenar por similitud descendente
        indices_match.sort(key=lambda j: scores[j], reverse=True)
        for j in indices_match:
            filas.append({
                "Practica Centro": practica_orig,
                "Valor":           centro.iloc[i]["valor"],
                "COD NBU":         nbu.iloc[j]["CODIGO"],
                "Descripcion NBU": nbu.iloc[j]["Determinaciones"],
                "Similitud":       round(scores[j], 4),
                "Estado":          "Encontrado",
            })

df_resultado = pd.DataFrame(filas)
df_resultado.to_csv("resultado_nbu.csv", sep=";", index=False, encoding="latin-1")

# --- Generar Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Busqueda NBU"

font_rojo = Font(color="FF0000", bold=True)
fill_rosa = PatternFill(fill_type="solid", fgColor="FFCCCC")

headers = df_resultado.columns.tolist()
ws.append(headers)

for _, row in df_resultado.iterrows():
    ws.append(row.tolist())
    if row["Estado"] == "No Encontrado":
        fila_excel = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=fila_excel, column=col)
            cell.font = font_rojo
            cell.fill = fill_rosa

wb.save("resultado_nbu.xlsx")

encontrados    = (df_resultado["Estado"] == "Encontrado").sum()
no_encontrados = (df_resultado["Estado"] == "No Encontrado").sum()
print(f"Encontrados:    {encontrados} coincidencias (umbral >= {UMBRAL})")
print(f"No encontrados: {no_encontrados} pr\u00e1cticas")
print("Archivos guardados: resultado_nbu.csv y resultado_nbu.xlsx")
